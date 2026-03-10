from __future__ import annotations

from typing import Dict

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

def write_report_excel(
    report_path: str,
    summary: Dict[str, float],
    unmapped_rows_df: pd.DataFrame,
    differences_df: pd.DataFrame,
    diff_matrix_df: pd.DataFrame,
) -> None:
    """
    Escribe un archivo Excel con:
      - Summary
      - Unmapped_Source_Rows
      - Differences (matriz DIFF para heatmap)
      - Differences_Detail (detalle por celda con error)

    En la hoja "Differences":
      - Cada celda numérica representa DIFF = EXPECTED - OUTPUT.
      - Se aplica un formato condicional:
          * Verde si DIFF = 0  (celda correcta)
          * Rojo si DIFF <> 0  (celda incorrecta)
    """
    summary_df = pd.DataFrame([summary])

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        # Summary
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Unmapped rows
        if not unmapped_rows_df.empty:
            unmapped_rows_df.to_excel(
                writer, sheet_name="Unmapped_Source_Rows", index=False
            )
        else:
            pd.DataFrame(columns=["INFO"]).to_excel(
                writer,
                sheet_name="Unmapped_Source_Rows",
                index=False,
            )

        # Matriz de diferencias (para heatmap)
        if not diff_matrix_df.empty:
            diff_matrix_df.to_excel(writer, sheet_name="Differences", index=False)
            wb = writer.book
            ws = writer.sheets["Differences"]

            # Determinar columnas de fecha (todas las que no son clave ni FILA)
            key_cols = [
                "COUNTRY",
                "CHANNEL",
                "NAMEPLATE",
                "TRIM",
                "CONCAT",
                "SEGMENT",
                "PARAMETER",
            ]
            all_cols = list(diff_matrix_df.columns)
            date_cols = [c for c in all_cols if c not in key_cols + ["FILA"]]

            if date_cols:
                # Indices 1-based de columnas de fecha
                date_col_indices = [all_cols.index(c) + 1 for c in date_cols]
                start_col_idx = min(date_col_indices)
                end_col_idx = max(date_col_indices)

                start_col_letter = get_column_letter(start_col_idx)
                end_col_letter = get_column_letter(end_col_idx)

                first_row = 2  # fila 1 = encabezados
                last_row = ws.max_row

                cell_range = f"{start_col_letter}{first_row}:{end_col_letter}{last_row}"

                green_fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                red_fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

                # Verde si = 0
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(operator="equal", formula=["0"], fill=green_fill),
                )
                # Rojo si <> 0
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(operator="notEqual", formula=["0"], fill=red_fill),
                )
        else:
            pd.DataFrame(columns=["INFO"]).to_excel(
                writer,
                sheet_name="Differences",
                index=False,
            )

        # Detalle de diferencias por celda (incluye FILA y COLUMN)
        if not differences_df.empty:
            differences_df.to_excel(
                writer, sheet_name="Differences_Detail", index=False
            )
        else:
            pd.DataFrame(columns=["INFO"]).to_excel(
                writer,
                sheet_name="Differences_Detail",
                index=False,
            )